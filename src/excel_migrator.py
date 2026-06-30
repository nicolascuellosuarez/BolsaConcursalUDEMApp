"""
Module to migrate extracted PDF data to Excel.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict, Optional
import os

from pdf_extractor import extract_pdf_data


# Excel headers (23 columns)
EXCEL_HEADERS = [
    "Solicitud",
    "Convocatoria Bolsa Concursal",
    "N. Comité",
    "Radicado de la solicitud",
    "Docente",
    "Programa",
    "Facultad",
    "Evento",
    "Fecha viaje",
    "Tipo de movilidad",
    "Destino",
    "Institución de destino",
    "Descripción del evento y la participación",
    "Monto solicitado",
    "Rubros",
    "Compromisos",
    "Observaciones",
    "Concepto del Comité de Movilidad",
    "Respuesta radicado",
    "Final solicitado en Abedul por la Facultad",
    "Evidencia Abedul",
    "Acta de compromisos",
    "Observaciones de la solicitud"
]


def find_last_row(worksheet) -> int:
    """Find the last row with data in the worksheet."""
    for row in range(worksheet.max_row, 0, -1):
        if any(worksheet.cell(row=row, column=col).value for col in range(1, worksheet.max_column + 1)):
            return row
    return 1


def auto_adjust_cells(worksheet, min_height: int = 40):
    """
    Adjust column widths and row heights based on content.
    """
    # Adjust column widths
    for col in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        
        for cell in col:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        
        adjusted_width = min(max(max_length + 5, 15), 80)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Adjust row heights
    for row in worksheet.iter_rows():
        max_lines = 1
        
        for cell in row:
            if cell.value:
                text = str(cell.value)
                col_width = worksheet.column_dimensions[get_column_letter(cell.column)].width or 50
                lines = len(text) // (col_width * 1.2) + text.count('\n') + 1
                if lines > max_lines:
                    max_lines = lines
        
        height = max(min_height, min(max_lines * 15, 200))
        worksheet.row_dimensions[row[0].row].height = height


def build_record(pdf_data: Dict[str, str], 
                 periodo: str, 
                 num_comite: str, 
                 fecha_comite: str,
                 num_solicitud: int) -> Dict[str, str]:
    """Build a complete Excel record from PDF data."""
    
    convocatoria = f"Convocatoria {periodo}"
    comite = f"Comité {num_comite} {fecha_comite}"
    solicitud_num = str(num_solicitud).zfill(3)
    
    return {
        "Solicitud": solicitud_num,
        "Convocatoria Bolsa Concursal": convocatoria,
        "N. Comité": comite,
        "Radicado de la solicitud": pdf_data.get("radicado", ""),
        "Docente": pdf_data.get("docente", ""),
        "Programa": "",
        "Facultad": pdf_data.get("facultad", ""),
        "Evento": pdf_data.get("evento", ""),
        "Fecha viaje": pdf_data.get("fechas_evento", ""),
        "Tipo de movilidad": pdf_data.get("tipo_movilidad", ""),
        "Destino": pdf_data.get("destino", ""),
        "Institución de destino": pdf_data.get("institucion", ""),
        "Descripción del evento y la participación": pdf_data.get("descripcion", ""),
        "Monto solicitado": pdf_data.get("monto_total", ""),
        "Rubros": pdf_data.get("rubros", ""),
        "Compromisos": pdf_data.get("compromisos", ""),
        "Observaciones": "",
        "Concepto del Comité de Movilidad": "",
        "Respuesta radicado": "",
        "Final solicitado en Abedul por la Facultad": "",
        "Evidencia Abedul": "",
        "Acta de compromisos": "",
        "Observaciones de la solicitud": ""
    }


def migrate_requests(excel_path: str, 
                     pdf_paths: List[str], 
                     periodo: str, 
                     num_comite: str, 
                     fecha_comite: str) -> Dict[str, any]:
    """
    Migrate PDF requests to Excel.
    """
    # Load Excel
    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        return {"error": f"File not found: {excel_path}"}
    
    # Select sheet by period
    sheet_name = "Solicitudes BC 2026 - 1" if periodo == "1" else "Solicitudes 2026 Año - 2"
    
    if sheet_name not in wb.sheetnames:
        return {"error": f"Sheet '{sheet_name}' not found in Excel"}
    
    ws = wb[sheet_name]
    
    # Find last row
    last_row = find_last_row(ws)
    next_row = last_row + 1 if last_row > 1 else 2
    
    # Get starting request number
    if last_row > 1:
        last_num = ws.cell(row=last_row, column=1).value
        try:
            start_num = int(last_num) + 1
        except (ValueError, TypeError):
            start_num = last_row
    else:
        start_num = 1
    
    # Process each PDF
    results = {
        "total": len(pdf_paths),
        "success": 0,
        "failed": 0,
        "errors": [],
        "records": []
    }
    
    for idx, pdf_path in enumerate(pdf_paths):
        try:
            pdf_data = extract_pdf_data(pdf_path)
            
            if not pdf_data:
                raise ValueError("No data extracted from PDF")
            
            solicitud_num = start_num + idx
            record = build_record(pdf_data, periodo, num_comite, fecha_comite, solicitud_num)
            
            for col_idx, header in enumerate(EXCEL_HEADERS, 1):
                value = record.get(header, "")
                cell = ws.cell(row=next_row + idx, column=col_idx, value=value)
                
                cell.font = Font(name="Calibri", size=11, bold=False)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            results["success"] += 1
            results["records"].append({
                "solicitud": record["Solicitud"],
                "docente": record["Docente"]
            })
            
            print(f"Added record: {record['Solicitud']} - {record['Docente']}")
            
        except Exception as e:
            results["failed"] += 1
            error_msg = f"Error in {pdf_path}: {str(e)}"
            results["errors"].append(error_msg)
            print(f"{error_msg}")
    
    # 👇 AJUSTAR CELDAS ANTES DE GUARDAR
    if results["success"] > 0:
        auto_adjust_cells(ws, min_height=45)
    
    wb.save(excel_path)
    
    return results


def print_results(results: Dict[str, any]):
    """Print migration results."""
    print("\n" + "="*60)
    print("MIGRATION RESULTS")
    print("="*60)
    print(f"Total PDFs: {results.get('total', 0)}")
    print(f"Success: {results.get('success', 0)}")
    print(f"Failed: {results.get('failed', 0)}")
    
    if results.get("errors"):
        print("\nErrors:")
        for error in results["errors"]:
            print(f"  - {error}")
    
    if results.get("records"):
        print("\nRecords added:")
        for rec in results["records"]:
            print(f"  - Request {rec['solicitud']}: {rec['docente']}")
    print("="*60)


if __name__ == "__main__":
    import os

    excel_path = "../Bolsa Concursal 2026 1, 2.xlsx"
    pdf_paths = [
        "../data/Ana Milena Montoya Ruiz.pdf"
    ]

    print("DIRECTORIO:", os.getcwd())
    print("Excel existe:", os.path.exists(excel_path))
    print("PDF existe:", os.path.exists(pdf_paths[0]))
    print("Cantidad PDFs:", len(pdf_paths))

    periodo = "1"
    num_comite = "1"
    fecha_comite = "15/06/2026"

    results = migrate_requests(
        excel_path,
        pdf_paths,
        periodo,
        num_comite,
        fecha_comite
    )

    print(results)