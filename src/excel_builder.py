from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

GREEN = "92d050"
FONT_HEADER = "Calibri"
FONT_HEADER_SIZE = 11
FONT_CONTENT = "Calibri"
FONT_CONTENT_SIZE = 11

HEADER_FOR_REQUESTS_PAGES = [
    "Solicitud",
    "Convocatoria Bolsa Concursal",
    "N. Comité",
    "Radicado de la solicitud",
    "Docente",
    "Programa",
    "Facultad",
    "Evento",
    "Fecha viaje",
    "Tipo de movilidad",
    "Destino",
    "Institución de destino",
    "Descripción del evento y la participación",
    "Monto solicitado",
    "Rubros",
    "Compromisos",
    "Observaciones",
    "Concepto del Comité de Movilidad",
    "Respuesta radicado",
    "Final solicitado en Abedul por la Facultad",
    "Evidencia Abedul",
    "Acta de compromisos",
    "Observaciones de la solicitud"
]

HEADER_FOR_LICENSES_PAGE = [
    "Solicitud",
    "Convocatoria Bolsa Concursal",
    "N. Comité",
    "Radicado de la solicitud",
    "Docente",
    "Programa",
    "Facultad",
    "Evento",
    "Fecha viaje",
    "Tipo de movilidad",
    "Destino",
    "Institución de destino",
    "Descripción del evento y la participación",
    "Monto solicitado",
    "Rubros",
    "Compromisos",
    "Observaciones",
    "Concepto del Comité de Movilidad",
    "Respuesta radicado",
    "Final solicitado en Abedul por la Facultad",
    "Evidencia Abedul",
    "Acta de compromisos",
    "Observaciones de la solicitud"
]

CONSOLIDATED_HEADER = [
    "Docente",
    "Monto Solicitado",
    "Facultad"
]

WIDTHS_23 = [
    12, 18, 12, 18, 25, 20, 18, 30, 15, 18, 20, 30, 50, 18, 40, 35, 30, 25, 20, 25, 18, 18, 30
]

def create_header_style():
    """
    Creating the Style of the Excel Header.
    """

    font = Font(name = FONT_HEADER, 
                size = FONT_HEADER_SIZE,
                bold = True)
    fill = PatternFill(start_color = GREEN,
                end_color = GREEN,
                fill_type = "solid")
    alignment = Alignment(horizontal = "center", 
                vertical = "center", 
                wrap_text = True)
    border = Border(left = Side(border_style = "thin", color = "000000"),
                right = Side(border_style = "thin", color = "000000"),
                top = Side(border_style = "thin", color = "000000"),
                bottom = Side(border_style = "thin", color = "000000"))

    return font, fill, alignment, border

def create_content_style():
    """
    Creating the Style of the Excel Content.
    """
    
    font = Font(name = FONT_CONTENT,
                size = FONT_CONTENT_SIZE,
                bold = False)
    alignment = Alignment(horizontal = "left",
                vertical = "center",
                wrap_text = True)
    border = Border(left = Side(border_style = "thin", color = "000000"),
                right = Side(border_style = "thin", color = "000000"),
                top = Side(border_style = "thin", color = "000000"),
                bottom = Side(border_style = "thin", color = "000000"))
    
    return font, alignment, border

def apply_style_to_header(cell):
    """
    Apply the Style to the Excel Header.
    """

    font, fill, alignment, border = create_header_style()

    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border

def apply_style_to_content(cell):
    """
    Apply the Style to the Excel Content.
    """

    font, alignment, border = create_content_style()

    cell.font = font
    cell.alignment = alignment
    cell.border = border

def adjust_column_width(worksheet, column = None):
    """
    Adjust the width of the columns in the Excel.
    """

    if column:
        for i, ancho in enumerate(column, 1):
            letter = get_column_letter(i)
            worksheet.column_dimensions[letter].width = ancho
    else:
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

                max_length = min(max_length, 57)
                worksheet.column_dimensions[column_letter].width = max_length + 5

def set_default_row_height(worksheet, height=35):
    """
    Establece altura por defecto para todas las filas.
    """
    
    worksheet.sheet_format.defaultRowHeight = height
    for row in range(1, worksheet.max_row + 200):  # +200 por si crecen
        worksheet.row_dimensions[row].height = height

def create_paid_licenses_page(workbook):
    """
    Create the 'Lic. Remunerada' Page in Excel.
    """

    worksheet = workbook.create_sheet("Lic. Remunerada")
    for column_idx, header in enumerate(HEADER_FOR_LICENSES_PAGE, 1):
        cell = worksheet.cell(row = 1, column = column_idx, value = header)
        apply_style_to_header(cell)

    set_default_row_height(worksheet, 40)
    adjust_column_width(worksheet, WIDTHS_23)
    return worksheet



def create_consolidated_page(workbook):
    """
    Creating the 'Consolidado' Page in Excel
    """

    worksheet = workbook.create_sheet("Consolidado")
    for column_idx, header in enumerate(CONSOLIDATED_HEADER, 1):
        cell = worksheet.cell(row = 1, column = column_idx, value = header)
        apply_style_to_header(cell)
    
    set_default_row_height(worksheet, 35)
    adjust_column_width(worksheet, [25, 20, 20])
    return worksheet



def create_requests_pages(workbook, name):
    """
    Create the 'Solicitudes' Page in Excel.
    """

    worksheet = workbook.create_sheet(name)
    for column_idx, header in enumerate(HEADER_FOR_REQUESTS_PAGES, 1):
        cell = worksheet.cell(row = 1, column = column_idx, value = header)
        apply_style_to_header(cell)

    widths = [
        12,  
        18, 
        12,  
        18, 
        25,  
        20,  
        18, 
        30,  
        15,  
        18, 
        20, 
        30,  
        50,  
        18, 
        40, 
        35,   
        30,  
        25,   
        20,  
        25,  
        18,
        18,  
        30   
    ]

    set_default_row_height(worksheet, 40)
    adjust_column_width(worksheet, widths)
    return worksheet



def create_empty_excel(route_to_save):
    """
    Creating Excel Archive with the 4 required sheets.
    """

    workbook = Workbook()
    workbook.remove(workbook.active)

    create_paid_licenses_page(workbook)
    create_consolidated_page(workbook)
    create_requests_pages(workbook, f"Solicitudes BC 2026 - 1")
    create_requests_pages(workbook, f"Solicitudes BC 2026 - 2")

    workbook.save(route_to_save)
    return route_to_save

